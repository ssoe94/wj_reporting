"""Deterministic intake for the quality team's OOXML workbooks.

The importer is intentionally an intake/review boundary.  It normalizes cells
and embedded pictures into draft records, but never creates ``QualityReport``.
"""

from __future__ import annotations

import hashlib
import json
import logging
import os
import posixpath
import re
import resource
import socket
import tempfile
import threading
import time as monotonic_time
import uuid
import warnings as python_warnings
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from decimal import Decimal, InvalidOperation
from io import BytesIO
from pathlib import PurePosixPath
from typing import Any, BinaryIO
from zipfile import BadZipFile, ZIP_DEFLATED, ZipFile, ZipInfo
from xml.etree import ElementTree

from django.conf import settings
from django.core.files.base import ContentFile
from django.db import IntegrityError, close_old_connections, connection, models, transaction
from django.utils import timezone
from openpyxl import load_workbook
from PIL import Image as PillowImage, ImageOps

from .models import (
    QualityImportAsset,
    QualityImportBatch,
    QualityImportMedia,
    QualityImportProvenance,
    QualityImportRow,
)
from .storage import quality_import_media_upload_available


XLSX_CONTENT_TYPES = {
    '',
    'application/octet-stream',
    'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
}
MAX_UPLOAD_BYTES = 80 * 1024 * 1024
MAX_ZIP_ENTRIES = 2048
MAX_UNCOMPRESSED_BYTES = 128 * 1024 * 1024
MAX_ZIP_ENTRY_BYTES = 64 * 1024 * 1024
MAX_COMPRESSION_RATIO = 100
MAX_ROWS_PER_SHEET = 5000
MAX_SUPPORTED_SHEETS = 16
MAX_TOTAL_ROWS = 10_000
MAX_NORMALIZED_ROW_BYTES = 32 * 1024 * 1024
MAX_MEDIA_ITEMS = 500
MAX_MEDIA_BYTES = 10_000_000
MAX_MEDIA_TOTAL_BYTES = 64 * 1024 * 1024
MAX_NORMALIZED_MEDIA_TOTAL_BYTES = 64 * 1024 * 1024
MAX_IMAGE_PIXELS = 10_000_000
NORMALIZED_IMAGE_LONG_EDGE = 1024
NORMALIZED_IMAGE_QUALITY = 90
NORMALIZER_VERSION = 'quality-image-v1'
MAX_CELL_TEXT = 20_000
MAX_XML_PART_BYTES = 16 * 1024 * 1024
MAX_STAGED_MEDIA_BYTES = 128 * 1024 * 1024

ISSUE_SHEET_PATTERN = re.compile(r'^\s*\d{1,2}\s*月\s*$')
OQC_SHEET_NAME = 'OQC出库不良 返工list'


class WorkbookValidationError(ValueError):
    def __init__(self, code: str, message: str):
        super().__init__(message)
        self.code = code
        self.message = message


@dataclass
class ParsedWorkbook:
    sheet_names: list[str]
    properties: dict[str, Any]
    rows: list[dict[str, Any]]
    media: list[dict[str, Any]]
    warnings: list[str]


@dataclass(frozen=True)
class ImportScope:
    """User-selected report-date scope for one workbook intake."""

    mode: str
    range_start: date | None = None
    range_end: date | None = None

    def as_dict(self) -> dict[str, str | None]:
        return {
            'mode': self.mode,
            'range_start': self.range_start.isoformat() if self.range_start else None,
            'range_end': self.range_end.isoformat() if self.range_end else None,
        }

    @property
    def key(self) -> str:
        if self.mode == 'full':
            return 'full'
        return f'{self.range_start.isoformat()}:{self.range_end.isoformat()}'


def parse_import_scope(mode: object, range_start: object, range_end: object) -> ImportScope:
    normalized_mode = str(mode or '').strip()
    if normalized_mode == 'full':
        if range_start or range_end:
            raise WorkbookValidationError(
                'unexpected_import_range',
                'Full-history import must not include range_start or range_end.',
            )
        return ImportScope(mode='full')
    if normalized_mode != 'date_range':
        raise WorkbookValidationError(
            'invalid_import_mode',
            'import_mode must be date_range or full.',
        )
    try:
        start = date.fromisoformat(str(range_start or ''))
        end = date.fromisoformat(str(range_end or ''))
    except ValueError as exc:
        raise WorkbookValidationError(
            'invalid_import_range',
            'range_start and range_end must use YYYY-MM-DD.',
        ) from exc
    if start > end:
        raise WorkbookValidationError(
            'invalid_import_range',
            'range_start must be on or before range_end.',
        )
    return ImportScope(mode='date_range', range_start=start, range_end=end)


def apply_import_scope(parsed: ParsedWorkbook, scope: ImportScope) -> tuple[ParsedWorkbook, dict[str, Any]]:
    """Select rows and their anchored photos without retaining excluded bytes."""

    source_total = len(parsed.rows)
    undated_rows = [row for row in parsed.rows if row.get('report_date') is None]
    undated = len(undated_rows)
    if scope.mode == 'full':
        selected_rows = list(parsed.rows)
        selected_dated_count = source_total - undated
    else:
        selected_dated_rows = [
            row for row in parsed.rows
            if row.get('report_date') is not None
            and scope.range_start <= row['report_date'] <= scope.range_end
        ]
        selected_dated_count = len(selected_dated_rows)
        # Undated incidents remain recoverable as a separate review bucket;
        # their existing missing_report_date warning prevents publication
        # until a reviewer supplies a date.
        selected_rows = selected_dated_rows + undated_rows
    selected_anchors = {
        (row['sheet_name'], row['source_row_number'])
        for row in selected_rows
    }
    selected_media = [
        item for item in parsed.media
        if (item['source_sheet_name'], item['source_anchor_row']) in selected_anchors
    ]
    selection = {
        **scope.as_dict(),
        'source_total_rows': source_total,
        'selected_rows': selected_dated_count,
        'retained_rows': len(selected_rows),
        'excluded_rows': source_total - len(selected_rows),
        'undated_rows': undated,
    }
    properties = dict(parsed.properties)
    properties['selection_scope'] = selection
    return ParsedWorkbook(
        sheet_names=list(parsed.sheet_names),
        properties=properties,
        rows=selected_rows,
        media=selected_media,
        warnings=list(parsed.warnings),
    ), selection


def sha256_bytes(content: bytes) -> str:
    return hashlib.sha256(content).hexdigest()


def validate_upload_metadata(filename: str, content_type: str, size: int) -> None:
    if not filename or not filename.lower().endswith('.xlsx'):
        raise WorkbookValidationError('invalid_extension', 'Only .xlsx workbooks are supported.')
    if content_type and content_type.lower() not in XLSX_CONTENT_TYPES:
        raise WorkbookValidationError('invalid_content_type', 'The uploaded file is not an XLSX workbook.')
    if size <= 0:
        raise WorkbookValidationError('empty_file', 'The uploaded workbook is empty.')
    if size > MAX_UPLOAD_BYTES:
        raise WorkbookValidationError(
            'file_too_large',
            f'Workbook exceeds the {MAX_UPLOAD_BYTES // (1024 * 1024)} MB upload limit.',
        )


def _validate_zip_member(info: ZipInfo) -> None:
    path = PurePosixPath(info.filename.replace('\\', '/'))
    if path.is_absolute() or '..' in path.parts:
        raise WorkbookValidationError('unsafe_zip_path', 'Workbook contains an unsafe archive path.')
    if info.flag_bits & 0x1:
        raise WorkbookValidationError('encrypted_workbook', 'Encrypted workbooks are not supported.')
    if info.file_size > MAX_ZIP_ENTRY_BYTES:
        raise WorkbookValidationError('zip_entry_too_large', 'Workbook contains an oversized archive entry.')
    if info.file_size and info.compress_size == 0:
        raise WorkbookValidationError('invalid_zip_entry', 'Workbook contains an invalid compressed entry.')
    if info.compress_size and info.file_size / info.compress_size > MAX_COMPRESSION_RATIO:
        raise WorkbookValidationError('unsafe_compression_ratio', 'Workbook compression ratio is unsafe.')


def _seekable_source(source: bytes | bytearray | BinaryIO) -> BinaryIO:
    if isinstance(source, (bytes, bytearray)):
        return BytesIO(bytes(source))
    if not hasattr(source, 'read') or not hasattr(source, 'seek'):
        raise WorkbookValidationError('invalid_ooxml', 'The workbook source is not seekable.')
    source.seek(0)
    return source


def validate_and_sanitize_ooxml(source: bytes | bytearray | BinaryIO) -> tuple[BinaryIO, list[str], list[str]]:
    """Validate OOXML without copying a potentially large uploaded workbook."""

    stream = _seekable_source(source)
    signature = stream.read(2)
    stream.seek(0)
    if signature != b'PK':
        raise WorkbookValidationError('invalid_ooxml', 'The file is not an OOXML ZIP container.')

    warnings: list[str] = []
    try:
        archive = ZipFile(stream, 'r')
    except BadZipFile as exc:
        raise WorkbookValidationError('invalid_ooxml', 'The XLSX archive is damaged.') from exc

    with archive:
        infos = archive.infolist()
        if len(infos) > MAX_ZIP_ENTRIES:
            raise WorkbookValidationError('too_many_zip_entries', 'Workbook contains too many archive entries.')
        total_size = 0
        names = set()
        for info in infos:
            _validate_zip_member(info)
            total_size += info.file_size
            if total_size > MAX_UNCOMPRESSED_BYTES:
                raise WorkbookValidationError('zip_too_large', 'Workbook expands beyond the safe limit.')
            names.add(info.filename)

        required = {'[Content_Types].xml', 'xl/workbook.xml'}
        if not required.issubset(names):
            raise WorkbookValidationError('invalid_ooxml_structure', 'Required XLSX parts are missing.')
        if any(name.startswith('xl/externalLinks/') for name in names):
            raise WorkbookValidationError('external_links_not_allowed', 'External workbook links are not allowed.')
        if 'xl/vbaProject.bin' in names or any(name.lower().endswith('.bin') for name in names):
            raise WorkbookValidationError('macros_not_allowed', 'Macro-enabled workbook content is not allowed.')

        content_types = archive.read('[Content_Types].xml')
        if b'spreadsheetml.sheet.main+xml' not in content_types:
            raise WorkbookValidationError('invalid_workbook_type', 'OOXML package is not a standard XLSX workbook.')

        formula_count = 0
        for info in infos:
            if info.filename.startswith('xl/worksheets/') and info.filename.endswith('.xml'):
                data = archive.read(info.filename)
                formula_count += len(re.findall(br'<f(?:\s|>)', data))
                if b'<autoFilter' in data and 'wps_filters_ignored' not in warnings:
                    warnings.append('wps_filters_ignored')

    if formula_count:
        warnings.append(f'formula_cells_ignored:{formula_count}')
    stream.seek(0)
    return stream, warnings, sorted(names)


def _json_value(value: Any) -> Any:
    if value is None or isinstance(value, (bool, int, float, str)):
        if isinstance(value, str) and len(value) > MAX_CELL_TEXT:
            return value[:MAX_CELL_TEXT]
        return value
    if isinstance(value, (date, datetime)):
        return value.isoformat()
    return str(value)[:MAX_CELL_TEXT]


def _text(value: Any, *, max_length: int | None = None) -> str:
    if value is None:
        return ''
    if isinstance(value, float) and value.is_integer():
        result = str(int(value))
    else:
        result = str(value)
    result = result.strip()
    limit = max_length or MAX_CELL_TEXT
    return result[:limit]


def _decimal_date_text(value: float) -> str:
    return format(Decimal(str(value)), 'f')


def _parse_date(
    value: Any,
    *,
    uploaded_on: date,
    explicit_year: int | None = None,
) -> tuple[date | None, list[str]]:
    if value in (None, ''):
        return None, ['missing_report_date']
    if isinstance(value, datetime):
        return value.date(), []
    if isinstance(value, date):
        return value, []

    text = _decimal_date_text(value) if isinstance(value, float) else _text(value)
    normalized = re.sub(r'[年/\-]', '.', text).replace('月', '.').replace('日', '')
    normalized = re.sub(r'\.+', '.', normalized).strip('.')
    parts = normalized.split('.')
    try:
        if len(parts) == 3:
            year = int(parts[0])
            year = year + 2000 if year < 100 else year
            return date(year, int(parts[1]), int(parts[2])), []
        if len(parts) == 2:
            month, day = int(parts[0]), int(parts[1])
            if explicit_year is not None:
                return date(explicit_year, month, day), []
            year = uploaded_on.year - 1 if month > uploaded_on.month + 1 else uploaded_on.year
            return date(year, month, day), [f'report_year_inferred:{year}']
    except (ValueError, TypeError):
        pass
    return None, [f'invalid_report_date:{text[:64]}']


def _positive_expression(value: str) -> int | None:
    clean = value.replace(',', '').strip()
    if not clean:
        return None
    match = re.match(r'^(\d+(?:\s*\+\s*\d+)*)', clean)
    if not match:
        return None
    try:
        numbers = [int(piece.strip()) for piece in match.group(1).split('+')]
    except ValueError:
        return None
    return sum(numbers)


def _pure_positive_expression(value: Any) -> int | None:
    """Parse only integers or integer addition, never unit-bearing prefixes."""

    clean = re.sub(r'\s+', '', _text(value).replace(',', ''))
    if not re.fullmatch(r'\d+(?:\+\d+)*', clean):
        return None
    return sum(int(piece) for piece in clean.split('+'))


def _explicit_year_from_oqc_title(ws, workbook_title: str = '') -> int | None:
    """Find a four-digit year in the OQC title row before using upload time."""

    title_values = [workbook_title]
    for values in ws.iter_rows(
        min_row=1,
        max_row=min(ws.max_row, 2),
        min_col=1,
        max_col=min(ws.max_column, 16),
        values_only=True,
    ):
        title_values.extend(_text(value) for value in values if value not in (None, ''))
    for value in title_values:
        match = re.search(r'(?<!\d)((?:19|20)\d{2})\s*年', _text(value))
        if match:
            return int(match.group(1))
    return None


def _parse_issue_quantities(value: Any) -> tuple[dict[str, Any], list[str]]:
    text = _text(value)
    result: dict[str, Any] = {
        'lot_qty': None,
        'inspection_qty': None,
        'defect_qty': None,
        'defect_rate': '',
    }
    warnings: list[str] = []
    if not text:
        return result, warnings

    lot_values = []
    for raw in re.findall(r'Lot\s*数\s*[:：]\s*([^\n\r]+)', text, flags=re.IGNORECASE):
        parsed = _positive_expression(raw)
        if parsed is not None:
            lot_values.append(parsed)
        if re.search(r'[^\d\s+,]', raw):
            warnings.append('lot_quantity_has_unit_or_note')
    if lot_values:
        result['lot_qty'] = sum(lot_values)
        if len(lot_values) > 1:
            warnings.append('multiple_lot_quantities_summed')

    defect_match = re.search(r'不良\s*数\s*[:：]\s*(\d[\d,]*)', text)
    if defect_match:
        result['defect_qty'] = int(defect_match.group(1).replace(',', ''))

    fraction = re.search(r'(\d[\d,]*)\s*/\s*(\d[\d,]*)\s*=\s*([\d.]+)\s*%', text)
    if fraction:
        result['defect_qty'] = int(fraction.group(1).replace(',', ''))
        result['inspection_qty'] = int(fraction.group(2).replace(',', ''))
        result['defect_rate'] = f'{fraction.group(3)}%'
    else:
        rate_match = re.search(r'不良\s*率\s*[:：]\s*([\d.]+)\s*%', text)
        if rate_match:
            result['defect_rate'] = f'{rate_match.group(1)}%'

    if not any(value not in (None, '') for value in result.values()):
        warnings.append('quantity_text_not_normalized')
    return result, warnings


def _number(value: Any) -> int | None:
    if value in (None, ''):
        return None
    if isinstance(value, bool):
        return None
    try:
        parsed = Decimal(str(value).replace(',', '').strip())
        if parsed < 0 or parsed != parsed.to_integral_value():
            return None
        return int(parsed)
    except (InvalidOperation, ValueError):
        return None


def _section_for_location(location: str) -> tuple[str, list[str]]:
    normalized = re.sub(r'\s+', '', location).upper()
    mapping = {
        '注塑': 'LQC_INJ',
        '射出': 'LQC_INJ',
        '加工': 'LQC_ASM',
        '组装': 'LQC_ASM',
        'IQC': 'IQC',
        'OQC': 'OQC',
        'LG': 'CS',
        'CS': 'CS',
    }
    if normalized in mapping:
        return mapping[normalized], []
    return '', [f'unmapped_occurrence_location:{location or "blank"}']


def _headers(values: tuple[Any, ...], first_col: int = 2) -> dict[int, str]:
    """Map zero-based tuple positions to non-empty header labels."""

    return {
        index: _text(value)
        for index, value in enumerate(values)
        if index + 1 >= first_col and _text(value)
    }


def _row_dict(values: tuple[Any, ...], headers: dict[int, str]) -> dict[str, Any]:
    return {
        header: _json_value(values[index] if index < len(values) else None)
        for index, header in headers.items()
    }


def _iter_rows_tolerating_wps_filter(values_iter):
    """Yield sheet data even when WPS emits an invalid trailing filter value.

    In OOXML, ``autoFilter`` follows ``sheetData``.  openpyxl's read-only parser
    therefore yields every row first and raises only while consuming the
    presentation filter.  Suppressing this one exact, already-audited error
    avoids a full in-memory workbook rewrite.
    """

    while True:
        try:
            yield next(values_iter)
        except StopIteration:
            return
        except ValueError as exc:
            if 'Value must be either numerical or a string containing a wildcard' in str(exc):
                return
            raise


def _business_key(
    *,
    kind: str,
    sheet_name: str,
    report_date: date | None,
    source_sequence: str,
    source_row_number: int,
    model: str,
    part_no: str,
    occurrence_location: str,
    phenomenon: str,
    fallback_content_sha256: str = '',
) -> str:
    identity = {
        'sheet_role': kind,
        'date': report_date.isoformat() if report_date else '',
    }
    if source_sequence:
        identity['sequence'] = source_sequence
    else:
        # Row numbers are presentation coordinates, not event identities: an
        # inserted spreadsheet row must never supersede a different event.
        # A source without an explicit sequence is only a conservative match
        # when the event tuple itself remains exactly equal.
        # Without an explicit business sequence, only an exact normalized
        # event match is safe across cumulative workbook revisions.  A looser
        # tuple can merge two legitimate incidents that differ in quantity,
        # item, disposition, or action result.
        identity['fallback_content_sha256'] = fallback_content_sha256
    return sha256_bytes(json.dumps(identity, ensure_ascii=False, sort_keys=True).encode())


def normalized_row_fingerprint(row: Any) -> str:
    def value(key):
        return row.get(key) if isinstance(row, dict) else getattr(row, key, None)

    identity = {
        key: value(key)
        for key in (
            'report_date', 'section', 'occurrence_location', 'model', 'part_no',
            'item_name', 'lot_qty', 'inspection_qty', 'defect_qty', 'defect_rate',
            'judgement', 'phenomenon', 'disposition', 'action_result',
        )
    }
    return sha256_bytes(json.dumps(identity, ensure_ascii=False, sort_keys=True, default=str).encode())


def _parse_issue_sheet(ws, uploaded_on: date, *, consume_row=None) -> list[dict[str, Any]]:
    if ws.max_row > MAX_ROWS_PER_SHEET:
        raise WorkbookValidationError('too_many_rows', f'{ws.title} exceeds the row limit.')
    row_limit = ws.max_row
    values_iter = ws.iter_rows(
        min_row=2,
        max_row=row_limit,
        min_col=1,
        max_col=min(ws.max_column, 32),
        values_only=True,
    )
    try:
        headers = _headers(tuple(next(values_iter)))
    except StopIteration as exc:
        raise WorkbookValidationError('invalid_issue_sheet', f'{ws.title} has no header row.') from exc
    required = {'序号', '发生日期', '发生场所', '不良现象'}
    if not required.issubset(set(headers.values())):
        raise WorkbookValidationError('invalid_issue_sheet', f'{ws.title} is missing required headers.')
    parsed: list[dict[str, Any]] = []
    for row_number, values in enumerate(_iter_rows_tolerating_wps_filter(values_iter), start=3):
        raw = _row_dict(tuple(values), headers)
        if not any(_text(raw.get(name)) for name in ('序号', '发生日期', '发生场所', 'Mold', 'P/N', '不良现象')):
            continue
        warnings: list[str] = []
        report_date, date_warnings = _parse_date(raw.get('发生日期'), uploaded_on=uploaded_on)
        warnings.extend(date_warnings)
        location = _text(raw.get('发生场所'), max_length=64)
        section, section_warnings = _section_for_location(location)
        warnings.extend(section_warnings)
        quantities, quantity_warnings = _parse_issue_quantities(raw.get('不良数量'))
        warnings.extend(quantity_warnings)
        part_no = _text(raw.get('P/N'), max_length=128).upper()
        if not part_no:
            warnings.append('missing_part_no')
        phenomenon = _text(raw.get('不良现象'))
        if not phenomenon:
            warnings.append('missing_phenomenon')
        item = _text(raw.get('Item'), max_length=255)
        source_sequence = _text(raw.get('序号'), max_length=64)
        row = {
            'sheet_name': ws.title,
            'sheet_role': 'monthly_issue',
            'source_row_number': row_number,
            'source_sequence': source_sequence,
            'report_date': report_date,
            'section': section,
            'occurrence_location': location,
            'model': _text(raw.get('Mold'), max_length=128),
            'part_no': part_no,
            'item_name': item,
            **quantities,
            'judgement': 'NG',
            'phenomenon': phenomenon,
            'disposition': '',
            'action_result': _text(raw.get('备注')),
            'raw_data': raw,
            'warnings': sorted(set(warnings)),
        }
        row['content_sha256'] = normalized_row_fingerprint(row)
        row['business_key'] = _business_key(
            kind='issue',
            sheet_name=ws.title,
            report_date=report_date,
            source_sequence=source_sequence,
            source_row_number=row_number,
            model=row['model'],
            part_no=row['part_no'],
            occurrence_location=row['occurrence_location'],
            phenomenon=row['phenomenon'],
            fallback_content_sha256=row['content_sha256'],
        )
        if consume_row:
            consume_row(row)
        parsed.append(row)
    return parsed


def _parse_oqc_sheet(
    ws,
    uploaded_on: date,
    *,
    workbook_title: str = '',
    consume_row=None,
) -> list[dict[str, Any]]:
    explicit_year = _explicit_year_from_oqc_title(ws, workbook_title)
    if ws.max_row > MAX_ROWS_PER_SHEET:
        raise WorkbookValidationError('too_many_rows', f'{ws.title} exceeds the row limit.')
    row_limit = ws.max_row
    values_iter = ws.iter_rows(
        min_row=2,
        max_row=row_limit,
        min_col=1,
        max_col=min(ws.max_column, 32),
        values_only=True,
    )
    try:
        headers = _headers(tuple(next(values_iter)))
    except StopIteration as exc:
        raise WorkbookValidationError('invalid_oqc_sheet', f'{ws.title} has no header row.') from exc
    required = {'检查日期', '型号', 'P/N', '不良类型'}
    if not required.issubset(set(headers.values())):
        raise WorkbookValidationError('invalid_oqc_sheet', f'{ws.title} is missing required headers.')
    parsed: list[dict[str, Any]] = []
    for row_number, values in enumerate(_iter_rows_tolerating_wps_filter(values_iter), start=3):
        raw = _row_dict(tuple(values), headers)
        if not any(_text(raw.get(name)) for name in ('检查日期', '型号', 'P/N', '不良类型', '数量')):
            continue
        warnings: list[str] = []
        report_date, date_warnings = _parse_date(
            raw.get('检查日期'),
            uploaded_on=uploaded_on,
            explicit_year=explicit_year,
        )
        warnings.extend(date_warnings)
        part_no = _text(raw.get('P/N'), max_length=128).upper()
        if not part_no:
            warnings.append('missing_part_no')
        phenomenon = _text(raw.get('不良类型'))
        if not phenomenon:
            warnings.append('missing_phenomenon')
        defect_qty = _pure_positive_expression(raw.get('数量'))
        if raw.get('数量') not in (None, '') and defect_qty is None:
            warnings.append('invalid_defect_quantity')
        source_sequence = _text(raw.get('NO'), max_length=64)
        occurrence_location = _text(raw.get('对应部门'), max_length=64)
        if not occurrence_location:
            warnings.append('missing_occurrence_location')
        row = {
            'sheet_name': ws.title,
            'sheet_role': 'oqc',
            'source_row_number': row_number,
            'source_sequence': source_sequence,
            'report_date': report_date,
            'section': 'OQC',
            'occurrence_location': occurrence_location,
            'model': _text(raw.get('型号'), max_length=128),
            'part_no': part_no,
            'item_name': _text(raw.get('品名'), max_length=255),
            'lot_qty': None,
            'inspection_qty': None,
            'defect_qty': defect_qty,
            'defect_rate': '',
            'judgement': 'NG',
            'phenomenon': phenomenon,
            'disposition': _text(raw.get('处理结果')),
            'action_result': _text(raw.get('备注')),
            'raw_data': raw,
            'warnings': sorted(set(warnings)),
        }
        row['content_sha256'] = normalized_row_fingerprint(row)
        row['business_key'] = _business_key(
            kind='oqc',
            sheet_name=ws.title,
            report_date=report_date,
            source_sequence=source_sequence,
            source_row_number=row_number,
            model=row['model'],
            part_no=row['part_no'],
            occurrence_location=row['occurrence_location'],
            phenomenon=row['phenomenon'],
            fallback_content_sha256=row['content_sha256'],
        )
        if consume_row:
            consume_row(row)
        parsed.append(row)
    return parsed


def _image_format(content: bytes) -> tuple[str, str] | None:
    if content.startswith(b'\x89PNG\r\n\x1a\n'):
        return 'png', 'image/png'
    if content.startswith(b'\xff\xd8\xff'):
        return 'jpg', 'image/jpeg'
    if content.startswith((b'GIF87a', b'GIF89a')):
        return 'gif', 'image/gif'
    if content.startswith(b'BM'):
        return 'bmp', 'image/bmp'
    if content.startswith((b'II*\x00', b'MM\x00*')):
        return 'tiff', 'image/tiff'
    if content.startswith(b'RIFF') and content[8:12] == b'WEBP':
        return 'webp', 'image/webp'
    return None


def _image_dimensions(content: bytes) -> tuple[int, int]:
    try:
        with python_warnings.catch_warnings():
            python_warnings.simplefilter('error', PillowImage.DecompressionBombWarning)
            with PillowImage.open(BytesIO(content)) as image:
                width, height = image.size
                frame_count = max(1, int(getattr(image, 'n_frames', 1)))
                if width <= 0 or height <= 0 or width * height * frame_count > MAX_IMAGE_PIXELS:
                    raise WorkbookValidationError(
                        'image_dimensions_too_large',
                        'An embedded image exceeds the safe pixel limit.',
                    )
                # ``open`` only parses the header. Verify the rest of the
                # encoded stream before accepting its dimensions or storing it.
                image.verify()
    except WorkbookValidationError:
        raise
    except (
        PillowImage.DecompressionBombError,
        PillowImage.DecompressionBombWarning,
        OSError,
        SyntaxError,
        ValueError,
    ) as exc:
        raise WorkbookValidationError(
            'invalid_embedded_image',
            'Workbook contains an unreadable or unsafe embedded image.',
        ) from exc
    return width, height


def _normalize_image_content(
    content: bytes,
    *,
    extension: str,
    content_type: str,
) -> tuple[bytes, str, str, int, int]:
    """Match the existing quality-photo policy deterministically on the server.

    The established browser path uses max long edge 1024, q=0.9, no upscale,
    and preserves the source format.  EXIF transpose is additionally enforced
    here before dimensions/deduplication so workbook pictures render correctly.
    """

    with PillowImage.open(BytesIO(content)) as source:
        source_format = (source.format or '').upper()
        image = ImageOps.exif_transpose(source)
        # Detach decoded pixels from the source container and explicitly drop
        # ICC/EXIF/XMP/text/density metadata. Pillow's PNG encoder otherwise
        # falls back to ``image.info['icc_profile']`` even without an explicit
        # encoder option, which would retain workbook-supplied metadata.
        image = image.copy()
        image.info.clear()
        needs_resize = max(image.size) > NORMALIZED_IMAGE_LONG_EDGE
        if needs_resize:
            image.thumbnail(
                (NORMALIZED_IMAGE_LONG_EDGE, NORMALIZED_IMAGE_LONG_EDGE),
                PillowImage.Resampling.LANCZOS,
            )
        output = BytesIO()
        if source_format in {'JPEG', 'JPG'}:
            if image.mode not in {'RGB', 'L'}:
                image = image.convert('RGB')
            image.save(output, format='JPEG', quality=NORMALIZED_IMAGE_QUALITY, optimize=True)
            normalized_extension, normalized_type = 'jpg', 'image/jpeg'
        elif source_format == 'PNG':
            image.save(output, format='PNG', optimize=True, icc_profile=None)
            normalized_extension, normalized_type = 'png', 'image/png'
        elif source_format == 'WEBP':
            image.save(
                output,
                format='WEBP',
                quality=NORMALIZED_IMAGE_QUALITY,
                method=6,
            )
            normalized_extension, normalized_type = 'webp', 'image/webp'
        else:
            # Canvas-based manual uploads flatten unsupported photo formats;
            # PNG is the deterministic, lossless safe fallback here.
            if image.mode not in {'RGB', 'RGBA', 'L', 'LA'}:
                image = image.convert('RGBA')
            image.save(output, format='PNG', optimize=True, icc_profile=None)
            normalized_extension, normalized_type = 'png', 'image/png'
        normalized = output.getvalue()
        if len(normalized) > MAX_MEDIA_BYTES:
            raise WorkbookValidationError(
                'normalized_image_too_large',
                'Normalized image exceeds the Cloudinary byte limit.',
            )
        return normalized, normalized_extension, normalized_type, image.width, image.height


def _safe_sheet_key(sheet_name: str) -> str:
    key = re.sub(r'[^0-9A-Za-z._-]+', '-', sheet_name).strip('-').lower()
    return key or 'sheet'


def _read_xml_bytes(archive: ZipFile, part: str) -> bytes:
    try:
        data = archive.read(part)
    except KeyError as exc:
        raise WorkbookValidationError('missing_ooxml_part', f'Missing OOXML part: {part}') from exc
    if len(data) > MAX_XML_PART_BYTES:
        raise WorkbookValidationError('xml_part_too_large', f'OOXML XML part is too large: {part}')
    upper = data.upper()
    if b'<!DOCTYPE' in upper or b'<!ENTITY' in upper:
        raise WorkbookValidationError('unsafe_xml_declaration', 'Workbook contains unsafe XML declarations.')
    return data


def _xml_root(archive: ZipFile, part: str):
    try:
        return ElementTree.fromstring(_read_xml_bytes(archive, part))
    except ElementTree.ParseError as exc:
        raise WorkbookValidationError('invalid_ooxml_xml', f'Invalid OOXML XML part: {part}') from exc


def _relationship_map(archive: ZipFile, relationship_path: str) -> dict[str, str]:
    if relationship_path not in archive.namelist():
        return {}
    root = _xml_root(archive, relationship_path)
    result = {}
    for relationship in root:
        target_mode = relationship.attrib.get('TargetMode', '')
        if target_mode and target_mode.lower() != 'internal':
            continue
        rel_id = relationship.attrib.get('Id')
        target = relationship.attrib.get('Target')
        if rel_id and target:
            result[rel_id] = target
    return result


def _resolve_part(base_part: str, target: str) -> str:
    if target.startswith('/'):
        return target.lstrip('/')
    return posixpath.normpath(posixpath.join(posixpath.dirname(base_part), target))


def _part_relationship_path(part: str) -> str:
    return posixpath.join(posixpath.dirname(part), '_rels', posixpath.basename(part) + '.rels')


def _workbook_sheet_parts(archive: ZipFile) -> list[tuple[str, str]]:
    main_ns = 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'
    rel_ns = 'http://schemas.openxmlformats.org/officeDocument/2006/relationships'
    workbook_part = 'xl/workbook.xml'
    relationships = _relationship_map(archive, 'xl/_rels/workbook.xml.rels')
    root = _xml_root(archive, workbook_part)
    result = []
    for sheet in root.findall(f'.//{{{main_ns}}}sheet'):
        rel_id = sheet.attrib.get(f'{{{rel_ns}}}id')
        target = relationships.get(rel_id or '')
        if target:
            resolved = _resolve_part(workbook_part, target)
            if resolved.startswith('xl/worksheets/'):
                result.append((sheet.attrib.get('name', ''), resolved))
    return result


def _parse_media(
    source: BinaryIO,
    workbook_sha256: str,
    *,
    heartbeat=None,
    allowed_anchors: set[tuple[str, int]] | None = None,
) -> tuple[list[dict[str, Any]], list[str]]:
    parsed: list[dict[str, Any]] = []
    warnings: list[str] = []
    total_bytes = 0
    normalized_total_bytes = 0
    main_ns = 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'
    rel_ns = 'http://schemas.openxmlformats.org/officeDocument/2006/relationships'
    drawing_ns = 'http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing'
    drawingml_ns = 'http://schemas.openxmlformats.org/drawingml/2006/main'

    source.seek(0)
    with ZipFile(source, 'r') as archive:
        names = set(archive.namelist())
        for sheet_name, sheet_part in _workbook_sheet_parts(archive):
            sheet_xml = _read_xml_bytes(archive, sheet_part)
            sheet_relationships = _relationship_map(archive, _part_relationship_path(sheet_part))
            drawing_refs = re.findall(
                br'<(?:[A-Za-z_][\w.-]*:)?drawing\b[^>]*\b(?:[A-Za-z_][\w.-]*:)?id=["\']([^"\']+)["\']',
                sheet_xml,
            )
            anchor_counts: dict[int, int] = {}
            for drawing_id_bytes in drawing_refs:
                drawing_id = drawing_id_bytes.decode('utf-8', errors='strict')
                drawing_target = sheet_relationships.get(drawing_id)
                if not drawing_target:
                    warnings.append(f'unreadable_drawing_reference:{sheet_name}')
                    continue
                drawing_part = _resolve_part(sheet_part, drawing_target)
                if not drawing_part.startswith('xl/drawings/') or drawing_part not in names:
                    warnings.append(f'missing_drawing_part:{sheet_name}')
                    continue
                drawing_relationships = _relationship_map(
                    archive,
                    _part_relationship_path(drawing_part),
                )
                drawing_root = _xml_root(archive, drawing_part)
                anchors = [
                    child for child in list(drawing_root)
                    if child.tag in {
                        f'{{{drawing_ns}}}twoCellAnchor',
                        f'{{{drawing_ns}}}oneCellAnchor',
                        f'{{{drawing_ns}}}absoluteAnchor',
                    }
                ]
                for anchor in anchors:
                    marker = anchor.find(f'{{{drawing_ns}}}from')
                    row_node = marker.find(f'{{{drawing_ns}}}row') if marker is not None else None
                    col_node = marker.find(f'{{{drawing_ns}}}col') if marker is not None else None
                    if row_node is None or col_node is None:
                        warnings.append(f'unreadable_image_anchor:{sheet_name}')
                        continue
                    row_number = int(row_node.text or '0') + 1
                    col_number = int(col_node.text or '0') + 1
                    if (
                        allowed_anchors is not None
                        and (sheet_name, row_number) not in allowed_anchors
                    ):
                        continue
                    if len(parsed) >= MAX_MEDIA_ITEMS:
                        raise WorkbookValidationError('too_many_images', 'Workbook exceeds the embedded image limit.')
                    blip = anchor.find(f'.//{{{drawingml_ns}}}blip')
                    embed_id = blip.attrib.get(f'{{{rel_ns}}}embed') if blip is not None else None
                    image_target = drawing_relationships.get(embed_id or '')
                    if not image_target:
                        warnings.append(f'unreadable_embedded_image:{sheet_name}:{row_number}')
                        continue
                    image_part = _resolve_part(drawing_part, image_target)
                    if not image_part.startswith('xl/media/') or image_part not in names:
                        warnings.append(f'missing_embedded_image:{sheet_name}:{row_number}')
                        continue
                    content_bytes = archive.read(image_part)
                    source_index = anchor_counts.get(row_number, 0)
                    anchor_counts[row_number] = source_index + 1
                    total_bytes += len(content_bytes)
                    if total_bytes > MAX_MEDIA_TOTAL_BYTES:
                        raise WorkbookValidationError('images_too_large', 'Embedded images exceed the aggregate safe limit.')
                    detected = _image_format(content_bytes)
                    if not detected:
                        warnings.append(f'unsupported_image_format:{sheet_name}:{row_number}:{source_index}')
                        continue
                    extension, content_type = detected
                    item_warnings = []
                    source_sha = sha256_bytes(content_bytes)
                    original_width = original_height = width = height = None
                    upload_content = None
                    normalized_extension = extension
                    normalized_content_type = content_type
                    try:
                        original_width, original_height = _image_dimensions(content_bytes)
                        (
                            upload_content,
                            normalized_extension,
                            normalized_content_type,
                            width,
                            height,
                        ) = _normalize_image_content(
                            content_bytes,
                            extension=extension,
                            content_type=content_type,
                        )
                    except WorkbookValidationError as exc:
                        item_warnings.append(exc.code)
                    normalized_sha = sha256_bytes(upload_content) if upload_content is not None else source_sha
                    if upload_content is not None:
                        normalized_total_bytes += len(upload_content)
                        if normalized_total_bytes > MAX_NORMALIZED_MEDIA_TOTAL_BYTES:
                            raise WorkbookValidationError(
                                'normalized_images_too_large',
                                'Normalized images exceed the aggregate safe limit.',
                            )
                    storage_key = f'quality-import/assets/{normalized_sha}'
                    parsed.append(
                        {
                            'source_sheet_name': sheet_name,
                            'source_anchor_row': row_number,
                            'source_anchor_col': col_number,
                            'source_index': source_index,
                            'original_filename': f'{_safe_sheet_key(sheet_name)}-r{row_number}-{source_index}.{extension}',
                            'content_type': normalized_content_type,
                            'byte_size': len(upload_content) if upload_content is not None else 0,
                            'sha256': normalized_sha,
                            'source_sha256': source_sha,
                            'original_byte_size': len(content_bytes),
                            'original_width': original_width,
                            'original_height': original_height,
                            'width': width,
                            'height': height,
                            'storage_key': storage_key,
                            'content': upload_content,
                            'extension': normalized_extension,
                            'warnings': item_warnings,
                        }
                    )
                    if heartbeat and len(parsed) % 25 == 0:
                        heartbeat()
    return parsed, warnings


def parse_quality_workbook(
    content: bytes | bytearray | BinaryIO,
    *,
    workbook_sha256: str,
    uploaded_on: date,
    heartbeat=None,
    import_scope: ImportScope | None = None,
) -> ParsedWorkbook:
    source, warnings, _names = validate_and_sanitize_ooxml(content)
    try:
        source.seek(0)
        workbook = load_workbook(
            source,
            data_only=True,
            read_only=True,
            keep_links=False,
        )
    except Exception as exc:
        raise WorkbookValidationError('unreadable_workbook', 'Workbook data could not be read safely.') from exc

    recognized = []
    rows: list[dict[str, Any]] = []
    normalized_row_bytes = 0
    normalized_row_count = 0

    def consume_row(row):
        nonlocal normalized_row_bytes, normalized_row_count
        normalized_row_count += 1
        if normalized_row_count > MAX_TOTAL_ROWS:
            raise WorkbookValidationError('too_many_total_rows', 'Workbook exceeds the total normalized row limit.')
        normalized_row_bytes += len(
            json.dumps(row, ensure_ascii=False, default=str).encode('utf-8')
        )
        if normalized_row_bytes > MAX_NORMALIZED_ROW_BYTES:
            raise WorkbookValidationError(
                'normalized_rows_too_large',
                'Workbook normalized row content exceeds the safe limit.',
            )
        if heartbeat and normalized_row_count % 100 == 0:
            heartbeat()
    workbook_title = _text(getattr(workbook.properties, 'title', ''))
    for ws in workbook.worksheets:
        is_oqc = ws.title == OQC_SHEET_NAME
        is_monthly_issue = bool(ISSUE_SHEET_PATTERN.match(ws.title))
        if not (is_oqc or is_monthly_issue):
            continue
        if len(recognized) >= MAX_SUPPORTED_SHEETS:
            raise WorkbookValidationError('too_many_supported_sheets', 'Workbook has too many supported sheets.')
        recognized.append(ws.title)
        if is_oqc:
            sheet_rows = _parse_oqc_sheet(
                ws,
                uploaded_on,
                workbook_title=workbook_title,
                consume_row=consume_row,
            )
        else:
            sheet_rows = _parse_issue_sheet(ws, uploaded_on, consume_row=consume_row)
        rows.extend(sheet_rows)
        if heartbeat:
            heartbeat()
    if not recognized:
        workbook.close()
        raise WorkbookValidationError('no_supported_sheets', 'No supported quality sheets were found.')

    properties = {
        'title': workbook_title,
        'creator': _text(getattr(workbook.properties, 'creator', '')),
        'created': _json_value(getattr(workbook.properties, 'created', None)),
        'modified': _json_value(getattr(workbook.properties, 'modified', None)),
        'recognized_sheets': recognized,
    }
    source_dataset_key, source_dataset_warnings = _dataset_period_from_rows(
        rows,
        fallback_date=uploaded_on,
    )
    properties['source_dataset_key'] = source_dataset_key
    properties['source_dataset_warnings'] = source_dataset_warnings
    all_sheet_names = list(workbook.sheetnames)
    workbook.close()
    allowed_anchors = None
    if import_scope is not None:
        scoped, _selection = apply_import_scope(
            ParsedWorkbook(
                sheet_names=all_sheet_names,
                properties=properties,
                rows=rows,
                media=[],
                warnings=warnings,
            ),
            import_scope,
        )
        rows = scoped.rows
        properties = scoped.properties
        if import_scope.mode == 'date_range':
            allowed_anchors = {
                (row['sheet_name'], row['source_row_number'])
                for row in rows
            }
    media, media_warnings = _parse_media(
        source,
        workbook_sha256,
        heartbeat=heartbeat,
        allowed_anchors=allowed_anchors,
    )
    warnings.extend(media_warnings)
    recognized_rows = {(row['sheet_name'], row['source_row_number']) for row in rows}
    for item in media:
        if (item['source_sheet_name'], item['source_anchor_row']) not in recognized_rows:
            item['warnings'].append('image_not_linked_to_import_row')
            item['content'] = None
            warnings.append(
                f'image_not_linked:{item["source_sheet_name"]}:{item["source_anchor_row"]}:{item["source_index"]}'
            )

    parsed = ParsedWorkbook(
        sheet_names=all_sheet_names,
        properties=properties,
        rows=rows,
        media=media,
        warnings=sorted(set(warnings)),
    )
    return parsed


MAX_PROCESS_ATTEMPTS = 3
LOGGER = logging.getLogger(__name__)


class QualityImportStop(RuntimeError):
    """Cooperative stop requested after a durable media checkpoint."""


def _dataset_period_from_rows(
    rows: list[dict[str, Any]],
    *,
    fallback_date: date,
) -> tuple[str, list[str]]:
    counts: dict[str, int] = {}
    for row in rows:
        report_date = row.get('report_date')
        if report_date:
            period = report_date.strftime('%Y-%m')
            counts[period] = counts.get(period, 0) + 1
    warnings: list[str] = []
    monthly_rows = [row for row in rows if row.get('sheet_role') == 'monthly_issue']
    explicit_months = {
        int(match.group(1))
        for row in monthly_rows
        if (match := re.fullmatch(r'\s*(\d{1,2})\s*月\s*', row.get('sheet_name', '')))
    }
    explicit_years: dict[int, int] = {}
    for row in monthly_rows:
        if row.get('report_date'):
            year = row['report_date'].year
            explicit_years[year] = explicit_years.get(year, 0) + 1
    if len(explicit_months) == 1 and explicit_years:
        month = next(iter(explicit_months))
        year = sorted(explicit_years, key=lambda value: (-explicit_years[value], value))[0]
        period = f'{year:04d}-{month:02d}'
    elif len(explicit_months) == 1:
        period = f'{fallback_date.year:04d}-{next(iter(explicit_months)):02d}'
        warnings.append('dataset_year_inferred_from_upload_date')
    elif counts:
        period = sorted(counts, key=lambda value: (-counts[value], value))[0]
    else:
        period = fallback_date.strftime('%Y-%m')
        warnings.append('dataset_period_inferred_from_upload_date')
    if len(counts) > 1:
        warnings.append('mixed_report_periods:' + ','.join(sorted(counts)))
    return f'quality_issue_workbook:{period}', warnings


def _dataset_period(parsed: ParsedWorkbook, batch: QualityImportBatch) -> tuple[str, list[str]]:
    source_key = parsed.properties.get('source_dataset_key')
    source_warnings = parsed.properties.get('source_dataset_warnings')
    if isinstance(source_key, str) and source_key.startswith('quality_issue_workbook:'):
        return source_key, list(source_warnings) if isinstance(source_warnings, list) else []
    return _dataset_period_from_rows(
        parsed.rows,
        fallback_date=timezone.localtime(batch.created_at).date(),
    )


def _row_evidence_sha(row: dict[str, Any], media: list[dict[str, Any]]) -> str:
    evidence = [
        {
            'column': item['source_anchor_col'],
            'index': item['source_index'],
            'sha256': item['sha256'],
        }
        for item in media
        if item['source_sheet_name'] == row['sheet_name']
        and item['source_anchor_row'] == row['source_row_number']
    ]
    return sha256_bytes(
        json.dumps(
            {'content': row['content_sha256'], 'media': sorted(evidence, key=lambda item: (item['column'], item['index']))},
            ensure_ascii=False,
            sort_keys=True,
        ).encode()
    )


def _make_delta_summary(rows: list[QualityImportRow], missing_rows: list[QualityImportRow]) -> dict[str, Any]:
    by_date: dict[str, dict[str, int]] = {}
    for row in rows:
        key = row.report_date.isoformat() if row.report_date else 'unknown'
        bucket = by_date.setdefault(key, {'added': 0, 'changed': 0, 'unchanged': 0, 'missing': 0})
        bucket[row.delta_status] += 1
    for row in missing_rows:
        key = row.report_date.isoformat() if row.report_date else 'unknown'
        by_date.setdefault(key, {'added': 0, 'changed': 0, 'unchanged': 0, 'missing': 0})['missing'] += 1
    return {'by_date': dict(sorted(by_date.items()))}


def _verify_stored_asset(storage, name: str, item: dict[str, Any]) -> bool:
    """Return False only when absent; fail closed for conflicting bytes."""

    if not name or not storage.exists(name):
        return False
    digest = hashlib.sha256()
    payload = bytearray()
    try:
        with storage.open(name, 'rb') as stored:
            while True:
                block = stored.read(1024 * 1024)
                if not block:
                    break
                payload.extend(block)
                if len(payload) > MAX_MEDIA_BYTES:
                    raise WorkbookValidationError(
                        'stored_asset_too_large',
                        'A normalized image in storage exceeds its safe byte limit.',
                    )
                digest.update(block)
    except WorkbookValidationError:
        raise
    except Exception as exc:
        raise WorkbookValidationError(
            'stored_asset_unavailable',
            'A normalized image could not be verified in durable storage.',
        ) from exc
    if len(payload) != item['byte_size'] or digest.hexdigest() != item['sha256']:
        raise WorkbookValidationError(
            'stored_asset_mismatch',
            'A content-addressed image failed checksum verification.',
        )
    detected = _image_format(bytes(payload))
    if not detected or detected[1] != item['content_type']:
        raise WorkbookValidationError(
            'stored_asset_type_mismatch',
            'A content-addressed image has an unexpected encoded type.',
        )
    width, height = _image_dimensions(bytes(payload))
    if width != item['width'] or height != item['height']:
        raise WorkbookValidationError(
            'stored_asset_dimensions_mismatch',
            'A content-addressed image has unexpected dimensions.',
        )
    return True


def _asset_item(asset: QualityImportAsset) -> dict[str, Any]:
    return {
        'sha256': asset.sha256,
        'byte_size': asset.byte_size,
        'content_type': asset.content_type,
        'width': asset.width,
        'height': asset.height,
    }


# ---------------------------------------------------------------------------
# Durable multipart intake
#
# Only normalized image bytes enter durable staging; XLSX bytes exist solely in
# an anonymous temporary file owned by the request and are closed before 202.
# ---------------------------------------------------------------------------

ASSET_LEASE_SECONDS = 5 * 60
REMOTE_VERIFICATION_TTL = timedelta(days=30)
PUMP_IDLE_GRACE_SECONDS = 0.25
_PUMP_GUARD = threading.Lock()
_PUMP_WAKE = threading.Event()
_PUMP_THREAD: threading.Thread | None = None
_STAGING_ADVISORY_LOCK = 0x57514C49  # "WQLI"


def _metadata_matches(asset: QualityImportAsset, item: dict[str, Any]) -> bool:
    return (
        asset.byte_size == item['byte_size']
        and asset.content_type == item['content_type']
        and asset.width == item['width']
        and asset.height == item['height']
        and asset.extension == item['extension']
        and asset.normalizer_version == NORMALIZER_VERSION
        and asset.storage_key == item['storage_key']
    )


def _lock_staging_capacity() -> None:
    if connection.vendor == 'postgresql':
        with connection.cursor() as cursor:
            cursor.execute('SELECT pg_advisory_xact_lock(%s)', [_STAGING_ADVISORY_LOCK])


def _pending_staged_bytes() -> int:
    return int(
        QualityImportAsset.objects.filter(staged_bytes__isnull=False).aggregate(
            total=models.Sum('byte_size')
        )['total']
        or 0
    )


def _latest_baseline_rows(
    *,
    dataset_key: str,
    selection_scope: dict[str, Any],
) -> tuple[QualityImportBatch | None, list[QualityImportRow], list[int]]:
    """Return the latest known row for each business key in this scope.

    A single latest batch is insufficient once daily date-scoped imports are
    allowed: yesterday's correction must still find yesterday's prior row even
    if today's batch is newer.  Missing calculations are scoped identically so
    rows outside the selected period are never reported as deleted.
    """

    queryset = QualityImportRow.objects.select_related('batch').filter(
        batch__dataset_key=dataset_key,
        batch__status__in=[
            QualityImportBatch.Status.READY,
            QualityImportBatch.Status.READY_WITH_WARNINGS,
        ],
    )
    if selection_scope.get('mode') == 'date_range':
        queryset = queryset.filter(
            models.Q(
                report_date__gte=selection_scope['range_start'],
                report_date__lte=selection_scope['range_end'],
            )
            | models.Q(report_date__isnull=True)
        )
    queryset = queryset.order_by('-batch__created_at', '-batch_id', '-id')
    latest_by_key: dict[str, QualityImportRow] = {}
    for row in queryset.iterator(chunk_size=500):
        latest_by_key.setdefault(row.business_key, row)
    rows = list(latest_by_key.values())
    batches = {row.batch_id: row.batch for row in rows}
    baseline_ids = sorted(batches)
    baseline = next(iter(batches.values())) if len(batches) == 1 else None
    return baseline, rows, baseline_ids


def _persist_parsed_workbook(
    *,
    parsed: ParsedWorkbook,
    source_sha256: str,
    source_filename: str,
    source_content_type: str,
    source_byte_size: int,
    uploaded_by,
    import_scope_key: str,
    selection_scope: dict[str, Any],
) -> tuple[QualityImportBatch, bool]:
    """Atomically persist review rows and normalized-image staging bytes."""

    with transaction.atomic():
        _lock_staging_capacity()
        existing = QualityImportBatch.objects.select_for_update().filter(
            sha256=source_sha256,
            import_scope_key=import_scope_key,
        ).first()
        if existing:
            return existing, True

        batch = QualityImportBatch.objects.create(
            uploaded_by=uploaded_by,
            original_filename=source_filename,
            sha256=source_sha256,
            import_scope_key=import_scope_key,
            file_size=source_byte_size,
            dataset_key='quality_issue_workbook:pending',
            status=QualityImportBatch.Status.QUEUED,
            phase='staging_results',
        )
        dataset_key, dataset_warnings = _dataset_period(parsed, batch)
        baseline, baseline_rows, baseline_batch_ids = _latest_baseline_rows(
            dataset_key=dataset_key,
            selection_scope=selection_scope,
        )
        candidates: dict[str, list[QualityImportRow]] = {}
        for baseline_row in baseline_rows:
            candidates.setdefault(baseline_row.business_key, []).append(baseline_row)

        media_by_row: dict[tuple[str, int], list[dict[str, Any]]] = {}
        for item in parsed.media:
            media_by_row.setdefault(
                (item['source_sheet_name'], item['source_anchor_row']),
                [],
            ).append(item)

        row_lookup: dict[tuple[str, int], QualityImportRow] = {}
        created_rows: list[QualityImportRow] = []
        consumed_baseline_ids: set[int] = set()
        for source_data in parsed.rows:
            data = dict(source_data)
            row_warnings = list(data.pop('warnings'))
            evidence_sha = _row_evidence_sha(
                data,
                media_by_row.get((data['sheet_name'], data['source_row_number']), []),
            )
            baseline_row = next(
                (
                    candidate
                    for candidate in candidates.get(data['business_key'], [])
                    if candidate.id not in consumed_baseline_ids
                ),
                None,
            )
            if baseline_row is None:
                delta_status = QualityImportRow.DeltaStatus.ADDED
                review_status = QualityImportRow.ReviewStatus.DRAFT
                supersedes = None
            elif (baseline_row.evidence_sha256 or baseline_row.content_sha256) == evidence_sha:
                consumed_baseline_ids.add(baseline_row.id)
                delta_status = QualityImportRow.DeltaStatus.UNCHANGED
                review_status = QualityImportRow.ReviewStatus.UNCHANGED
                supersedes = None
            else:
                consumed_baseline_ids.add(baseline_row.id)
                delta_status = QualityImportRow.DeltaStatus.CHANGED
                review_status = QualityImportRow.ReviewStatus.DRAFT
                supersedes = baseline_row
                row_warnings.append(f'source_row_revision_of:{baseline_row.id}')
            source_key = sha256_bytes(
                f'{source_sha256}:{data["sheet_name"]}:{data["source_row_number"]}'.encode()
            )
            row = QualityImportRow.objects.create(
                batch=batch,
                source_key=source_key,
                evidence_sha256=evidence_sha,
                baseline_row=baseline_row,
                supersedes=supersedes,
                delta_status=delta_status,
                review_status=review_status,
                warnings=sorted(set(row_warnings)),
                **data,
            )
            created_rows.append(row)
            row_lookup[(row.sheet_name, row.source_row_number)] = row

        missing_rows = [row for row in baseline_rows if row.id not in consumed_baseline_ids]
        staged_asset_cache: dict[str, QualityImportAsset] = {}
        new_asset_shas: set[str] = set()
        reused_attachments = 0
        media_warning_count = 0
        staged_bytes_before = _pending_staged_bytes()
        staged_bytes_added = 0
        verification_cutoff = timezone.now() - REMOTE_VERIFICATION_TTL
        verification_required_by_sha: dict[str, bool] = {}
        for item in parsed.media:
            if item.get('content') is None:
                continue
            linked_row = row_lookup.get(
                (item['source_sheet_name'], item['source_anchor_row'])
            )
            requires_verification = (
                linked_row is None
                or linked_row.delta_status != QualityImportRow.DeltaStatus.UNCHANGED
            )
            verification_required_by_sha[item['sha256']] = (
                verification_required_by_sha.get(item['sha256'], False)
                or requires_verification
            )

        for item in parsed.media:
            asset = None
            content = item.get('content')
            if content is not None:
                asset = staged_asset_cache.get(item['sha256'])
                if asset is None:
                    asset = QualityImportAsset.objects.select_for_update().filter(
                        sha256=item['sha256']
                    ).first()
                    if asset:
                        metadata_matches = _metadata_matches(asset, item)
                        active_upload = (
                            asset.upload_state == QualityImportAsset.UploadState.UPLOADING
                            and bool(asset.processing_owner)
                            and bool(asset.lease_expires_at)
                            and asset.lease_expires_at > timezone.now()
                        )
                        reusable_unchanged_asset = (
                            not verification_required_by_sha[item['sha256']]
                            and metadata_matches
                            and asset.upload_state == QualityImportAsset.UploadState.READY
                            and bool(asset.file and asset.file.name)
                            and bool(asset.remote_verified_at)
                            and asset.remote_verified_at >= verification_cutoff
                        )
                        if not reusable_unchanged_asset and not active_upload:
                            if asset.staged_bytes is None:
                                staged_bytes_added += item['byte_size']
                                if staged_bytes_before + staged_bytes_added > MAX_STAGED_MEDIA_BYTES:
                                    raise WorkbookValidationError(
                                        'staging_capacity_exceeded',
                                        'Pending normalized images exceed the temporary database staging limit.',
                                    )
                            # ADDED/CHANGED evidence and any stale or incomplete
                            # unchanged checkpoint must be verified by the pump.
                            # Metadata is deterministic from these normalized
                            # bytes, so an inconsistent row is repaired here.
                            asset.normalizer_version = NORMALIZER_VERSION
                            asset.byte_size = item['byte_size']
                            asset.content_type = item['content_type']
                            asset.width = item['width']
                            asset.height = item['height']
                            asset.extension = item['extension']
                            asset.storage_key = item['storage_key']
                            asset.staged_bytes = content
                            asset.upload_state = QualityImportAsset.UploadState.STAGED
                            asset.processing_owner = ''
                            asset.lease_expires_at = None
                            asset.next_attempt_at = None
                            asset.last_error = ''
                            asset.save(update_fields=[
                                'normalizer_version', 'byte_size', 'content_type',
                                'width', 'height', 'extension', 'storage_key',
                                'staged_bytes', 'upload_state', 'processing_owner',
                                'lease_expires_at', 'next_attempt_at', 'last_error',
                            ])
                    else:
                        staged_bytes_added += item['byte_size']
                        if staged_bytes_before + staged_bytes_added > MAX_STAGED_MEDIA_BYTES:
                            raise WorkbookValidationError(
                                'staging_capacity_exceeded',
                                'Pending normalized images exceed the temporary database staging limit.',
                            )
                        asset = QualityImportAsset.objects.create(
                            sha256=item['sha256'],
                            normalizer_version=NORMALIZER_VERSION,
                            byte_size=item['byte_size'],
                            content_type=item['content_type'],
                            width=item['width'],
                            height=item['height'],
                            extension=item['extension'],
                            storage_key=item['storage_key'],
                            file='',
                            staged_bytes=content,
                            upload_state=QualityImportAsset.UploadState.STAGED,
                            created_by_batch=batch,
                        )
                        new_asset_shas.add(item['sha256'])
                    staged_asset_cache[item['sha256']] = asset
                if item['sha256'] not in new_asset_shas:
                    reused_attachments += 1

            media_warnings = list(item.get('warnings') or [])
            media_warning_count += len(media_warnings)
            QualityImportMedia.objects.create(
                batch=batch,
                row=row_lookup.get((item['source_sheet_name'], item['source_anchor_row'])),
                asset=asset,
                source_sheet_name=item['source_sheet_name'],
                source_anchor_row=item['source_anchor_row'],
                source_anchor_col=item['source_anchor_col'],
                source_index=item['source_index'],
                original_filename=item['original_filename'],
                source_sha256=item['source_sha256'],
                source_byte_size=item['original_byte_size'],
                source_width=item['original_width'],
                source_height=item['original_height'],
                warnings=media_warnings,
            )

        all_warnings = sorted(set(parsed.warnings + dataset_warnings))
        warning_count = (
            len(all_warnings)
            + sum(len(row.warnings) for row in created_rows)
            + media_warning_count
        )
        added = sum(row.delta_status == QualityImportRow.DeltaStatus.ADDED for row in created_rows)
        changed = sum(row.delta_status == QualityImportRow.DeltaStatus.CHANGED for row in created_rows)
        unchanged = sum(row.delta_status == QualityImportRow.DeltaStatus.UNCHANGED for row in created_rows)
        pending_assets = {
            asset.pk
            for asset in staged_asset_cache.values()
            if asset.upload_state != QualityImportAsset.UploadState.READY
        }
        now = timezone.now()
        QualityImportProvenance.objects.create(
            batch=batch,
            source_sha256=source_sha256,
            source_content_type=source_content_type,
            source_filename=source_filename,
            source_byte_size=source_byte_size,
            workbook_properties=parsed.properties,
            source_discarded_at=now,
        )
        ready_status = (
            QualityImportBatch.Status.READY_WITH_WARNINGS
            if warning_count
            else QualityImportBatch.Status.READY
        )
        batch.dataset_key = dataset_key
        batch.baseline_batch = baseline
        batch.sheet_names = parsed.sheet_names
        batch.total_rows = len(parsed.rows)
        batch.total_media = len(parsed.media)
        batch.source_total_rows = int(selection_scope['source_total_rows'])
        batch.added_count = added
        batch.changed_count = changed
        batch.unchanged_count = unchanged
        batch.missing_count = len(missing_rows)
        batch.new_media_count = len(new_asset_shas)
        batch.reused_media_count = reused_attachments
        batch.delta_summary = {
            **_make_delta_summary(created_rows, missing_rows),
            'selection_scope': selection_scope,
            'baseline_batch_ids': baseline_batch_ids,
        }
        batch.warnings = all_warnings
        batch.warning_count = warning_count
        batch.progress_done = len(staged_asset_cache) - len(pending_assets)
        batch.progress_total = len(staged_asset_cache)
        batch.results_persisted_at = now
        batch.status = QualityImportBatch.Status.QUEUED if pending_assets else ready_status
        batch.phase = 'uploading_media' if pending_assets else 'ready'
        batch.save()
        return batch, False


def ingest_quality_workbook(
    upload,
    *,
    uploaded_by,
    import_scope: ImportScope,
) -> tuple[QualityImportBatch, bool]:
    """Stream one multipart upload, parse it, and discard its raw bytes."""

    filename = str(getattr(upload, 'name', '') or '')
    content_type = str(getattr(upload, 'content_type', '') or '')
    validate_upload_metadata(filename, content_type, int(getattr(upload, 'size', 0) or 0))
    if not quality_import_media_upload_available():
        raise WorkbookValidationError(
            'production_storage_required',
            'Cloudinary image storage is required for quality imports in production.',
        )

    digest = hashlib.sha256()
    size = 0
    parsed = None
    source_sha256 = ''
    with tempfile.TemporaryFile() as source:
        try:
            for block in upload.chunks(chunk_size=1024 * 1024):
                if not block:
                    continue
                size += len(block)
                if size > MAX_UPLOAD_BYTES:
                    raise WorkbookValidationError(
                        'file_too_large',
                        f'Workbook exceeds the {MAX_UPLOAD_BYTES // (1024 * 1024)} MiB upload limit.',
                    )
                digest.update(block)
                source.write(block)
            validate_upload_metadata(filename, content_type, size)
            source_sha256 = digest.hexdigest()
            existing = QualityImportBatch.objects.filter(
                sha256=source_sha256,
                import_scope_key=import_scope.key,
            ).first()
            if existing:
                return existing, True
            source.flush()
            source.seek(0)
            parsed = parse_quality_workbook(
                source,
                workbook_sha256=source_sha256,
                uploaded_on=timezone.localdate(),
                import_scope=import_scope,
            )
            selection_scope = parsed.properties['selection_scope']
            if import_scope.mode == 'date_range' and not parsed.rows:
                raise WorkbookValidationError(
                    'no_rows_in_selected_range',
                    'No quality rows were found in the selected date range.',
                )
            source.seek(0, os.SEEK_END)
        finally:
            try:
                upload.close()
            except Exception:
                pass

    if parsed is None:
        raise WorkbookValidationError('unreadable_workbook', 'Workbook parsing did not complete.')
    batch, replay = _persist_parsed_workbook(
        parsed=parsed,
        source_sha256=source_sha256,
        source_filename=filename,
        source_content_type=content_type,
        source_byte_size=size,
        uploaded_by=uploaded_by,
        import_scope_key=import_scope.key,
        selection_scope=selection_scope,
    )
    if batch.status == QualityImportBatch.Status.QUEUED:
        transaction.on_commit(kick_quality_import_pump)
    return batch, replay


def recover_stale_quality_imports() -> int:
    """Return expired upload leases to the durable DB queue."""

    now = timezone.now()
    recovered = 0
    with transaction.atomic():
        stale = list(
            QualityImportBatch.objects.select_for_update().filter(
                status=QualityImportBatch.Status.PROCESSING,
                lease_expires_at__lt=now,
            )
        )
        for batch in stale:
            QualityImportAsset.objects.filter(
                attachments__batch=batch,
                upload_state=QualityImportAsset.UploadState.UPLOADING,
                lease_expires_at__lt=now,
            ).update(
                upload_state=QualityImportAsset.UploadState.STAGED,
                processing_owner='',
                lease_expires_at=None,
            )
            warnings = list(batch.warnings or [])
            if 'stale_processing_lease_recovered' not in warnings:
                warnings.append('stale_processing_lease_recovered')
            exhausted = batch.attempt_count >= MAX_PROCESS_ATTEMPTS
            batch.status = QualityImportBatch.Status.FAILED if exhausted else QualityImportBatch.Status.QUEUED
            batch.phase = 'failed' if exhausted else 'retry_wait'
            batch.processing_owner = ''
            batch.lease_expires_at = None
            batch.next_attempt_at = None if exhausted else now
            batch.warnings = warnings
            batch.warning_count = max(batch.warning_count, len(warnings))
            batch.save(update_fields=[
                'status', 'phase', 'processing_owner', 'lease_expires_at',
                'next_attempt_at', 'warnings', 'warning_count', 'updated_at',
            ])
            recovered += 1
    return recovered


def claim_quality_import_batch(batch_id: int | None = None) -> tuple[QualityImportBatch, str] | None:
    recover_stale_quality_imports()
    now = timezone.now()
    try:
        with transaction.atomic():
            queryset = QualityImportBatch.objects.select_for_update().filter(
                status=QualityImportBatch.Status.QUEUED,
            ).filter(
                models.Q(next_attempt_at__isnull=True) | models.Q(next_attempt_at__lte=now)
            ).order_by('created_at', 'id')
            if batch_id is not None:
                queryset = queryset.filter(pk=batch_id)
            batch = queryset.first()
            if batch is None:
                return None
            owner = f'{socket.gethostname()}:{os.getpid()}:{uuid.uuid4().hex}'
            batch.status = QualityImportBatch.Status.PROCESSING
            batch.phase = 'uploading_media'
            batch.processing_owner = owner
            batch.lease_expires_at = now + timedelta(seconds=ASSET_LEASE_SECONDS)
            batch.last_heartbeat_at = now
            batch.attempt_count += 1
            batch.next_attempt_at = None
            batch.save(update_fields=[
                'status', 'phase', 'processing_owner', 'lease_expires_at',
                'last_heartbeat_at', 'attempt_count', 'next_attempt_at', 'updated_at',
            ])
            return batch, owner
    except IntegrityError:
        # The conditional unique constraint permits one processing batch across
        # every web process. A competing pump simply backs off.
        return None


def _renew_upload_lease(batch_id: int, owner: str, *, done: int, total: int) -> None:
    now = timezone.now()
    updated = QualityImportBatch.objects.filter(
        pk=batch_id,
        status=QualityImportBatch.Status.PROCESSING,
        processing_owner=owner,
    ).update(
        phase='uploading_media',
        progress_done=max(0, done),
        progress_total=max(0, total),
        last_heartbeat_at=now,
        lease_expires_at=now + timedelta(seconds=ASSET_LEASE_SECONDS),
    )
    if not updated:
        raise WorkbookValidationError('processing_lease_lost', 'The media upload lease was lost.')


def _upload_staged_asset(asset_id: int, *, owner: str) -> QualityImportAsset:
    with transaction.atomic():
        asset = QualityImportAsset.objects.select_for_update().get(pk=asset_id)
        if asset.upload_state == QualityImportAsset.UploadState.READY:
            return asset
        if asset.staged_bytes is None:
            raise WorkbookValidationError(
                'normalized_checkpoint_incomplete',
                'A pending image has no normalized staging bytes.',
            )
        now = timezone.now()
        if (
            asset.upload_state == QualityImportAsset.UploadState.UPLOADING
            and asset.processing_owner
            and asset.processing_owner != owner
            and asset.lease_expires_at
            and asset.lease_expires_at > now
        ):
            raise WorkbookValidationError('asset_lease_busy', 'The image is already being uploaded.')
        asset.upload_state = QualityImportAsset.UploadState.UPLOADING
        asset.processing_owner = owner
        asset.lease_expires_at = now + timedelta(seconds=ASSET_LEASE_SECONDS)
        asset.attempt_count += 1
        asset.next_attempt_at = None
        asset.last_error = ''
        content = bytes(asset.staged_bytes)
        item = _asset_item(asset)
        requested_name = asset.storage_key
        asset.save(update_fields=[
            'upload_state', 'processing_owner', 'lease_expires_at',
            'attempt_count', 'next_attempt_at', 'last_error',
        ])

    if sha256_bytes(content) != asset.sha256 or len(content) != asset.byte_size:
        raise WorkbookValidationError(
            'staged_asset_mismatch',
            'Normalized staging bytes failed their content-address checksum.',
        )
    storage = QualityImportAsset._meta.get_field('file').storage
    existing_name = asset.file.name if asset.file and asset.file.name else ''
    if _verify_stored_asset(storage, existing_name, item):
        saved_name = existing_name
    elif requested_name != existing_name and _verify_stored_asset(storage, requested_name, item):
        saved_name = requested_name
    else:
        saved_name = storage.save(requested_name, ContentFile(content))
        if not _verify_stored_asset(storage, saved_name, item):
            raise WorkbookValidationError(
                'stored_asset_unavailable',
                'A normalized image could not be verified in durable storage.',
            )

    with transaction.atomic():
        asset = QualityImportAsset.objects.select_for_update().get(pk=asset_id)
        if asset.upload_state == QualityImportAsset.UploadState.READY:
            return asset
        if asset.processing_owner != owner:
            raise WorkbookValidationError('asset_lease_lost', 'The image upload lease was lost.')
        asset.file = saved_name
        asset.staged_bytes = None
        asset.upload_state = QualityImportAsset.UploadState.READY
        asset.processing_owner = ''
        asset.lease_expires_at = None
        asset.next_attempt_at = None
        asset.last_error = ''
        asset.remote_verified_at = timezone.now()
        asset.save(update_fields=[
            'file', 'staged_bytes', 'upload_state', 'processing_owner',
            'lease_expires_at', 'next_attempt_at', 'last_error', 'remote_verified_at',
        ])
    return asset


def _finish_batch(batch_id: int, owner: str) -> QualityImportBatch:
    with transaction.atomic():
        batch = QualityImportBatch.objects.select_for_update().get(pk=batch_id)
        if batch.processing_owner != owner or batch.status != QualityImportBatch.Status.PROCESSING:
            raise WorkbookValidationError('processing_lease_lost', 'The media upload lease was lost.')
        pending = batch.media.filter(asset__isnull=False).exclude(
            asset__upload_state=QualityImportAsset.UploadState.READY
        ).exists()
        invalid_anchor = batch.media.filter(asset__isnull=True, warnings=[]).exists()
        if pending or invalid_anchor:
            raise WorkbookValidationError(
                'normalized_checkpoint_incomplete',
                'Not every image attachment reached a durable checkpoint.',
            )
        batch.status = (
            QualityImportBatch.Status.READY_WITH_WARNINGS
            if batch.warning_count
            else QualityImportBatch.Status.READY
        )
        batch.phase = 'ready'
        batch.progress_done = batch.progress_total
        batch.processing_owner = ''
        batch.lease_expires_at = None
        batch.next_attempt_at = None
        batch.last_heartbeat_at = timezone.now()
        batch.save(update_fields=[
            'status', 'phase', 'progress_done', 'processing_owner',
            'lease_expires_at', 'next_attempt_at', 'last_heartbeat_at', 'updated_at',
        ])
    return QualityImportBatch.objects.get(pk=batch_id)


def _record_processing_failure(batch_id: int, owner: str, exc: Exception) -> None:
    code = exc.code if isinstance(exc, WorkbookValidationError) else 'media_upload_failed'
    message = exc.message if isinstance(exc, WorkbookValidationError) else str(exc)[:320]
    with transaction.atomic():
        batch = QualityImportBatch.objects.select_for_update().get(pk=batch_id)
        if batch.processing_owner != owner:
            return
        retryable = batch.attempt_count < MAX_PROCESS_ATTEMPTS
        now = timezone.now()
        QualityImportAsset.objects.filter(
            attachments__batch=batch,
            upload_state=QualityImportAsset.UploadState.UPLOADING,
            processing_owner=owner,
        ).update(
            upload_state=(
                QualityImportAsset.UploadState.STAGED
                if retryable
                else QualityImportAsset.UploadState.FAILED
            ),
            processing_owner='',
            lease_expires_at=None,
            next_attempt_at=(
                now + timedelta(minutes=min(60, 2 ** max(0, batch.attempt_count - 1)))
                if retryable
                else None
            ),
            last_error=f'{code}:{message}'[:512],
        )
        warning = f'{code}:{message}'[:1024]
        warnings = list(batch.warnings or [])
        if warning not in warnings:
            warnings.append(warning)
        batch.status = QualityImportBatch.Status.QUEUED if retryable else QualityImportBatch.Status.FAILED
        batch.phase = 'retry_wait' if retryable else 'failed'
        batch.processing_owner = ''
        batch.lease_expires_at = None
        batch.next_attempt_at = (
            now + timedelta(minutes=min(60, 2 ** max(0, batch.attempt_count - 1)))
            if retryable
            else None
        )
        batch.warnings = warnings
        batch.warning_count = max(batch.warning_count, len(warnings))
        batch.save(update_fields=[
            'status', 'phase', 'processing_owner', 'lease_expires_at',
            'next_attempt_at', 'warnings', 'warning_count', 'updated_at',
        ])


def process_quality_import_batch(batch_id: int | None = None, *, should_stop=None) -> QualityImportBatch | None:
    """Upload normalized staged images for one durable batch."""

    started_at = monotonic_time.monotonic()
    claimed = claim_quality_import_batch(batch_id)
    if claimed is None:
        return None
    batch, owner = claimed
    try:
        from .incremental_import import (  # pylint: disable=import-outside-toplevel
            INCREMENTAL_JOB_DATASET_KEY,
            finalize_quality_import_job,
        )
        if (
            batch.dataset_key == INCREMENTAL_JOB_DATASET_KEY
            and batch.total_media
            and not quality_import_media_upload_available()
        ):
            raise WorkbookValidationError(
                'production_storage_required',
                'Cloudinary image storage is required for Excel quality imports in production.',
            )
        asset_ids = list(
            QualityImportAsset.objects.filter(attachments__batch=batch)
            .exclude(upload_state=QualityImportAsset.UploadState.READY)
            .order_by('id')
            .values_list('id', flat=True)
            .distinct()
        )
        total = batch.media.filter(asset__isnull=False).values('asset_id').distinct().count()
        ready_before = max(0, total - len(asset_ids))
        _renew_upload_lease(batch.id, owner, done=ready_before, total=total)
        for offset, asset_id in enumerate(asset_ids, start=1):
            if should_stop and should_stop():
                raise QualityImportStop('Processing stop requested.')
            _upload_staged_asset(asset_id, owner=owner)
            _renew_upload_lease(batch.id, owner, done=ready_before + offset, total=total)
        # Incremental direct-import jobs reuse this durable asset queue, then
        # create reports only after every image reached remote storage. Import
        # lazily to keep the workbook parser dependency one-directional.
        if batch.dataset_key == INCREMENTAL_JOB_DATASET_KEY:
            finalize_quality_import_job(batch.id, owner)
        result = _finish_batch(batch.id, owner)
        LOGGER.info(
            'Quality import batch %s completed status=%s rows=%s media=%s elapsed_seconds=%.3f max_rss=%s',
            result.pk,
            result.status,
            result.total_rows,
            result.total_media,
            monotonic_time.monotonic() - started_at,
            resource.getrusage(resource.RUSAGE_SELF).ru_maxrss,
        )
        return result
    except QualityImportStop:
        QualityImportBatch.objects.filter(pk=batch.pk, processing_owner=owner).update(
            status=QualityImportBatch.Status.QUEUED,
            phase='queued',
            processing_owner='',
            lease_expires_at=None,
            next_attempt_at=timezone.now(),
        )
        return None
    except Exception as exc:
        _record_processing_failure(batch.id, owner, exc)
        LOGGER.exception('Quality import batch %s media upload failed', batch.id)
        raise


def retry_quality_import_batch(batch: QualityImportBatch) -> QualityImportBatch:
    with transaction.atomic():
        batch = QualityImportBatch.objects.select_for_update().get(pk=batch.pk)
        if batch.status != QualityImportBatch.Status.FAILED:
            return batch
        QualityImportAsset.objects.filter(
            attachments__batch=batch,
            upload_state=QualityImportAsset.UploadState.FAILED,
            staged_bytes__isnull=False,
        ).update(
            upload_state=QualityImportAsset.UploadState.STAGED,
            processing_owner='',
            lease_expires_at=None,
            next_attempt_at=None,
            last_error='',
        )
        if batch.media.filter(
            asset__upload_state=QualityImportAsset.UploadState.FAILED,
            asset__staged_bytes__isnull=True,
        ).exists():
            raise WorkbookValidationError(
                'source_reupload_required',
                'A failed normalized image no longer has retry bytes.',
            )
        batch.status = QualityImportBatch.Status.QUEUED
        batch.phase = 'queued'
        batch.processing_owner = ''
        batch.lease_expires_at = None
        batch.next_attempt_at = None
        batch.attempt_count = 0
        batch.save(update_fields=[
            'status', 'phase', 'processing_owner', 'lease_expires_at',
            'next_attempt_at', 'attempt_count', 'updated_at',
        ])
    transaction.on_commit(kick_quality_import_pump)
    return batch


def _seconds_until_next_queued_batch() -> float | None:
    next_attempt = QualityImportBatch.objects.filter(
        status=QualityImportBatch.Status.QUEUED,
        next_attempt_at__isnull=False,
    ).aggregate(next=models.Min('next_attempt_at'))['next']
    if next_attempt is None:
        return None
    return max(0.0, (next_attempt - timezone.now()).total_seconds())


def _pump_main() -> None:
    global _PUMP_THREAD
    close_old_connections()
    try:
        while True:
            _PUMP_WAKE.clear()
            while True:
                try:
                    result = process_quality_import_batch()
                except Exception:
                    result = None
                finally:
                    close_old_connections()
                if result is None:
                    break
            retry_delay = _seconds_until_next_queued_batch()
            if retry_delay is not None:
                if _PUMP_WAKE.wait(min(max(retry_delay, 0.05), 60.0)):
                    continue
                continue
            if _PUMP_WAKE.wait(PUMP_IDLE_GRACE_SECONDS):
                continue
            with _PUMP_GUARD:
                if _PUMP_WAKE.is_set():
                    continue
                _PUMP_THREAD = None
                return
    finally:
        close_old_connections()
        with _PUMP_GUARD:
            if _PUMP_THREAD is threading.current_thread():
                _PUMP_THREAD = None


def kick_quality_import_pump() -> None:
    """Lazily start one daemon in this process; DB leases arbitrate processes."""

    global _PUMP_THREAD
    configured = getattr(settings, 'QUALITY_IMPORT_DISABLE_BACKGROUND_PUMP', None)
    if configured is None:
        configured = os.getenv('QUALITY_IMPORT_DISABLE_BACKGROUND_PUMP', '')
    if isinstance(configured, str):
        configured = configured.strip().lower() in {'1', 'true', 'yes', 'on'}
    if bool(configured):
        return
    _PUMP_WAKE.set()
    with _PUMP_GUARD:
        if _PUMP_THREAD and _PUMP_THREAD.is_alive():
            return
        _PUMP_THREAD = threading.Thread(
            target=_pump_main,
            name='quality-import-media-pump',
            daemon=True,
        )
        _PUMP_THREAD.start()
